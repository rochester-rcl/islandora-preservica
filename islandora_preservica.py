import os
import os.path
import shutil
import pathlib
import re
import xml.etree.ElementTree as ET
from datetime import datetime
from bdbag import bdbag_api
from bagit import BagValidationError
from bdbag.bdbagit import BaggingInterruptedError
from pyrsistent import thaw
from zipfile import ZipFile
from openpyxl import Workbook
from os.path import basename

#------------------------------------------------------------------------------------------------------------------------------------------------------
# WORKFLOW FOR PREPARING ASSETS FOR PRESERVICA INGEST
# Assumes that access assets are coming from Islandora
# Assumes that preservation assets are coming from Digital Scholarship
# Assumes that metadata is coming from Islandora
#------------------------------------------------------------------------------------------------------------------------------------------------------
# ------manual process---------------------------------------------------------------------------------------------------------------------------------
# Manual Process - COPY preservations masters directory into root of project folder
# create_container() - Reanme 'ds_files' directory into container directory which will be dumped into WinSCP for OPEX incremental ingest
# folder_ds_files() - Transform single directory of preservation images into separate subdirectories full of images per asset
# create_bags_dir() -  Create bags directory to stage exported bags for processing
# Manual Process - COPY zipped bags over into created bags directory
# extract_bags() - Extract/unzip the bags in the bags directory
# validate_bags() - Validate the unzipped bags to ensure no errors in transfer
# create_id_ss() - Create a spreadsheet with the mapping between preservation file names, access file names, and bag ids
# Manual Process - Rectify the mismatches presented in the pres_acc_bag_ids spreadsheet
# ------can be run in sequnce--------------------------------------------------------------------------------------------------------------------------
# representation_preservation() - Create 'Representation_Preservation' subdirectories in each asset folder, then move preservation assets into them
# cleanup_metadata() - Removes excess XML headers from individual metadata files
# rename_bags() - Renames the bag directories to remove the "Bag-" prefix (if needed)
# process_bags() - Reverts the bags into simple directories and removes unnecessary files in 'data' subdirectory
# representation_access() - Create 'Representation_Access' subdirectories in each asset folder
# access_id_path() - Generate file containing MODS identifier and relative paths
# merge_access_preservation() - Loop through dirs in container, move access copies and metadata into relevant folders
# cleanup_bags() - Delete the bags_dir folder and the access_ids.txt file once merge is complete
# pax_metadata() - Write the OPEX metadata for the individua assets contained in the PAX
# stage_pax_content() - moves the Representation_Access and Representation_Preservation folders into a staging directory to enable zipping the PAX
# create_pax() - Make a PAX zip archive out of the Representation_Access and Representation_Access
# cleanup_directories() - Delete the xml files used to create the OPEX metadata and the directories used to create the PAX zip archive
# ao_opex_metadata() - Create the OPEX metadata for the archival object folder that syncs with ArchivesSpace, and rename subdirectories
# write_opex_container_md() - Write the OPEX metadata for the entire container structure
#------------------------------------------------------------------------------------------------------------------------------------------------------
# Project Log File variables by index
# 0 - date_time
# 1 - container
# 2 - bags_dir
#------------------------------------------------------------------------------------------------------------------------------------------------------

#NOTE orig_dir is folder of preservation masters, subdir of project folder
orig_dir = 'preservation_masters'
#NOTE proj_path may need to be updated based on OS of development environment
proj_path = 'M:/IDT/DAM/Perkins-Gillman_Ingest'
proj_log_file = os.path.join(proj_path, 'project_log.txt')

#this function takes the folder containing all the preservation masters and renames to be the "container" folder which will ultimately be used for OPEX incremental ingest
#also creates a "project_log.txt" file to store variables so that an ingest project can be worked on over multiple sessions
def create_container():
    project_log_hand = open(proj_log_file, 'a')
    now = datetime.now()
    date_time = now.strftime('%Y-%m-%d_%H-%M-%S')
    project_log_hand.write(date_time + '\n')
    container = 'container_' + date_time
    os.rename(os.path.join(proj_path, orig_dir), os.path.join(proj_path, container))
    project_log_hand.write(container + '\n')
    print('Container directory: {}'.format(container))
    project_log_hand.close()

#this function takes all of the preservation master files that come from DS in one big directory and splits them up into subdirectories containing all the images for a
#particular resource
def folder_ds_files():
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    project_log_hand.close()
    container = vars[1].strip()
    folder_name = ''
    folder_count = 0
    file_count = 0
    path_container = os.path.join(proj_path, container)
    path_foldername = os.path.join(proj_path, container, folder_name)
    for file in os.listdir(path = path_container):
        path_containerfile = os.path.join(proj_path, container, file)
        path_foldernamefile = os.path.join(proj_path, container, folder_name, file)
        if file.startswith('bags_'):
            continue
        else:
            file_root = file.split('-')[0]
            if  file_root == folder_name:
                shutil.move(path_containerfile, path_foldernamefile)
                file_count += 1
            else:
                folder_name = file_root
                os.mkdir(path_foldername)
                folder_count += 1
                shutil.move(path_containerfile, path_foldernamefile)
                file_count += 1
    for folder in os.listdir(path = path_container):
        path_folder = os.path.join(proj_path, container, folder)
        if folder.startswith('bags_'):
            continue
        else:
            num_files = len(os.listdir(path_folder))
            if num_files > 99:
                folder_name = os.path.join(proj_path, container, folder + "-001-" + str(num_files))
                os.rename(path_folder, folder_name)
            elif num_files > 9:
                folder_name = os.path.join(proj_path, container, folder + "-001-0" + str(num_files))
                os.rename(path_folder, folder_name)
            else:
                folder_name = os.path.join(proj_path, container, folder + "-001-00" + str(num_files))
                os.rename(path_folder, folder_name)
            print('{} created'.format(folder_name))
    print('Created and renamed {} subdirectories and moved {} files into them'.format(folder_count, file_count))

#this function creates a subdir in "container" to hold all the bags exported from Islandora, and manipulate them separate from the preservation masters
def create_bags_dir():
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    date_time = vars[0].strip()
    container = vars[1].strip()
    bags_dir = 'bags_' + date_time
    os.mkdir(os.path.join(proj_path, container, bags_dir))
    project_log_hand.close()
    project_log_hand = open(proj_log_file, 'a')
    project_log_hand.write(bags_dir + '\n')
    print('Created bags directory: {}'.format(bags_dir))
    project_log_hand.close()

#this function extracts the zipped bags into unzipped bags, and then deletes the zipped bags
def extract_bags():
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    container = vars[1].strip()
    bags_dir = vars[2].strip()
    num_bags = 0
    path_bagsdir = os.path.join(proj_path, container, bags_dir)
    for file in os.listdir(path = path_bagsdir):
        path_bagsdirfile = os.path.join(proj_path, container, bags_dir, file)
        bdbag_api.extract_bag(path_bagsdirfile, output_path = path_bagsdir, temp=False)
        print('extracting bag: {}'.format(file))
        num_bags += 1
    for bag in os.listdir(path = path_bagsdir):
        path_bagsdirbag = os.path.join(proj_path, container, bags_dir, bag)
        if bag.endswith('.zip'):
            print('removing zipped bag: {}'.format(bag))
            os.remove(path_bagsdirbag)
    print('Extracted {} bags'.format(str(num_bags)))
    project_log_hand.close()

#this function validates the bags to ensure the checksums don't indicate any corruption of files and checks for any other types of erros
#also logs the errors to a "validation_error_log.txt" for a record of problems which is also used in a later function
def validate_bags():
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    project_log_hand.close()
    container = vars[1].strip()
    bags_dir = vars[2].strip()
    error_log_handle = open(os.path.join(proj_path, 'validation_error_log.txt'), 'a')
    num_bags = 0
    path_bagsdir = os.path.join(proj_path, container, bags_dir)
    for directory in os.listdir(path = path_bagsdir):
        path_directory = os.path.join(proj_path, container, bags_dir, directory)
        print('attempting to validate bag: {}'.format(directory))
        num_bags += 1
        try:
            bdbag_api.validate_bag(path_directory, fast = False)
        except BagValidationError:
            error_log_handle.write('Bag Validation Error | Directory: ' + directory + '\n')
        except BaggingInterruptedError:
            error_log_handle.write('Bagging Interruped Error | Directory: ' + directory + '\n')
        except RuntimeError:
            error_log_handle.write('Runtime Error | Directory: ' + directory + '\n')
    print('Validated {} bags'.format(str(num_bags)))
    error_log_handle.close()

#this function creates an Excel spreadsheet that attemtps to match preservation assets and access assets up to each other
#will likely uncover preservation assets with no access corrolaries and vice versa which will require manual rectification
def create_id_ss():
    wb = Workbook()
    ws = wb.active
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    container = vars[1].strip()
    bags_dir = vars[2].strip()
    project_log_hand.close()
    ws['A1'] = 'pres_file_name'
    ws['B1'] = 'acc_file_name'
    ws['C1'] = 'bag_id'
    pres_file_list = []
    path_container = os.path.join(proj_path, container)
    path_bagsdir = os.path.join(proj_path, container, bags_dir)
    for folder in os.listdir(path = path_container):
        if folder.startswith('bags_'):
            continue
        else:
            pres_file_list.append(folder)
    bag_dict = dict()
    for bag in os.listdir(path =  path_bagsdir):
        path_bagmd = os.path.join(proj_path, container, bags_dir, bag, 'MODS.xml')
        tree = ET.parse(path_bagmd)
        identifier = tree.find('{http://www.loc.gov/mods/v3}identifier').text
        bag_dict[identifier] = bag
    for item in pres_file_list:
        if item in bag_dict.keys():
            ws.append([item, item, bag_dict[item]])
        else:
            ws.append([item, '', ''])
    for item in bag_dict.keys():
        if item not in pres_file_list:
            ws.append(['', item, bag_dict[item]])
    wb.save('pres_acc_bag_ids_suppl.xlsx')
    print('Created pres_acc_bag_ids.xlsx')

#------------------------------------------------------------------------------------------------------------------------

#this function begins the process of creating the PAX structure necessary for ingest
#"Representation_Preservation" folder is created, and each image is given a separate subdir inside of it
def representation_preservation():
    print('----CREATING REPRESENTATION_PRESERVATION FOLDERS AND MOVING ASSETS INTO THEM----')
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    project_log_hand.close()
    container = vars[1].strip()
    folder_count = 0
    file_count = 0
    path_container = os.path.join(proj_path, container)
    rep_pres = 'Representation_Preservation'
    for directory in os.listdir(path = path_container):
        path_directory = os.path.join(proj_path, container, directory)
        if directory.startswith('bags_'):
            continue
        path = os.path.join(proj_path, container, directory, rep_pres)
        os.mkdir(path)
        folder_count += 1
        for file in os.listdir(path = path_directory):
            path_directoryfile = os.path.join(proj_path, container, directory, file)
            if file == rep_pres:
                continue
            else:
                file_name = file.split('.')[0]
                os.mkdir(os.path.join(path, file_name))
                print('created directory: {}'.format(path + '/' + file_name))
                shutil.move(path_directoryfile, os.path.join(path, file_name, file))
                print('moved file: {}'.format(path + '/' + file_name + '/' + file))
            file_count += 1
    print('Created {} Representation_Preservation directories | Moved {} files into created directories'.format(folder_count, file_count))
    
# renames the bags if necessary depending on project needs
def rename_bags():
    print('----RENAMING BAGS----')
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    project_log_hand.close()
    container = vars[1].strip()
    num_bags = 0
    path_bagsdir = os.path.join(proj_path, container)
    for directory in os.listdir(path = path_bagsdir):
        path_bagsdirdirectory = os.path.join(proj_path, container, directory)
        id_name = directory.split('-')[1].strip()
        os.rename(path_bagsdirdirectory, os.path.join(path_bagsdir, id_name))
        num_bags += 1
        print('renamed {} into {}'.format(directory, id_name))
    print('renamed {} bags'.format(num_bags))

#this function processes the access assets and metadata contained within the Islandora bags and reverts the bag into a simple directory without the bag manifests
#the function renames the access asset by checking the MODS record and pulling the title field
#this function removes many unnecessary files provided by Islandora during bag export, ultimately leaving the access asset and any metadata files
def process_bags():
    print('----PROCESSING BAGS----')
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    project_log_hand.close()
    container = vars[1].strip()
    bags_dir = vars[2].strip()
    num_bags = 0
    error_log_handle = open(os.path.join(proj_path, 'validation_error_log.txt'), 'r')
    error_log = error_log_handle.read()
    error_log_handle.close()
    error_log_str = ''
    for line in error_log:
        error_log_str = error_log_str + line
    path_bagsdir = os.path.join(proj_path, container, bags_dir)
    for directory in os.listdir(path = path_bagsdir):
        path_bagsdirdirectory = os.path.join(proj_path, container, bags_dir, directory)
        #skips any directories that raised errors during validation
        if error_log_str.find(directory) != -1 :
            continue
        else:
            print('attempting to revert bag: {}'.format(directory))
            obj_file_name = ''
            extension = ''
            #converts the bags back into normal directories, removing bagit and manifest files
            bdbag_api.revert_bag(path_bagsdirdirectory)
            #removes unnecessary files generated by Islandora
            unnecessary_files = ['foo.xml', 'foxml.xml', 'JP2.jp2', 'JPG.jpg', 'POLICY.xml', 'PREVIEW.jpg', 'RELS-EXT.rdf', 'RELS-INT.rdf', 'TN.jpg', 'HOCR.html', 'OCR.txt', 'MP4.mp4', 'PROXY_MP3.mp3', 'TIFF.tif']
            for file in os.listdir(path = path_bagsdirdirectory):
                if file in unnecessary_files:
                    os.remove(os.path.join(proj_path, container, bags_dir, directory, file))
                if re.search('^OBJ', file):
                    obj_file_name = file
                    extension = obj_file_name.split('.')[1].strip()
                elif re.search('^PDF', file):
                    obj_file_name = file
                    extension = obj_file_name.split('.')[1].strip()
            path_objfilename = os.path.join(proj_path, container, bags_dir, directory, obj_file_name)
            #use xml.etree to identify filename from MODS.xml
            tree = ET.parse(os.path.join(path_bagsdirdirectory, 'MODS.xml'))
            identifier = tree.find('{http://www.loc.gov/mods/v3}identifier').text
            #rename the OBJ file to original filename pulled from MODS.xml
            os.rename(path_objfilename, os.path.join(path_bagsdirdirectory, identifier + '.' + extension))
        num_bags += 1
    print('Processed {} bags'.format(str(num_bags)))
    
#possible alternative to process_bags(), reverting the bags as a separate function
def revert_bags():
    print('----RERVERTING BAGS----')
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    project_log_hand.close()
    container = vars[1].strip()
    bags_dir = vars[2].strip()
    num_bags = 0
    error_log_handle = open(os.path.join(proj_path, 'validation_error_log.txt'), 'r')
    error_log = error_log_handle.read()
    error_log_handle.close()
    error_log_str = ''
    for line in error_log:
        error_log_str = error_log_str + line
    path_bagsdir = os.path.join(proj_path, container, bags_dir)
    for directory in os.listdir(path = path_bagsdir):
        path_bagsdirdirectory = os.path.join(proj_path, container, bags_dir, directory)
        #skips any directories that raised errors during validation
        if error_log_str.find(directory) != -1 :
            continue
        else:
            print('attempting to revert bag: {}'.format(directory))
            #converts the bags back into normal directories, removing bagit and manifest files
            bdbag_api.revert_bag(path_bagsdirdirectory)
            num_bags += 1
    print('Reverted {} bags'.format(str(num_bags)))
    
#possible alternative to process_bags() in cases where both preservation and access assets are exported from Islandora
#uses REGEX to identify a number of different file format extensions
def process_bags_islandora():
    print('----PROCESSING BAGS----')
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    project_log_hand.close()
    container = vars[1].strip()
    num_bags = 0
    path_bagsdir = os.path.join(proj_path, container)
    for directory in os.listdir(path = path_bagsdir):
        path_bagsdirdirectory = os.path.join(proj_path, container, directory)
        #skips any directories that raised errors during validation
        print('processing: {}'.format(directory))
        obj_file_name = ''
        extension = ''
        #removes unnecessary files generated by Islandora
        unnecessary_files = ['foo.xml', 'foxml.xml', 'JP2.jp2', 'JPG.jpg', 'POLICY.xml', 'PREVIEW.jpg', 'RELS-EXT.rdf', 'RELS-INT.rdf', 'TN.jpg', 'HOCR.html', 'OCR.txt', 'PROXY_MP3.mp3', 'TIFF.tif']
        for file in os.listdir(path = path_bagsdirdirectory):
            if file in unnecessary_files:
                os.remove(os.path.join(proj_path, container, directory, file))
        for file in os.listdir(path = path_bagsdirdirectory):
            if re.search('^OBJ', file):
                obj_file_name = file
                extension = obj_file_name.split('.')[1].strip()
                path_objfilename = os.path.join(proj_path, container, directory, obj_file_name)
                #use xml.etree to identify filename from MODS.xml
                tree = ET.parse(os.path.join(path_bagsdirdirectory, 'DC.xml'))
                identifier = tree.find('{http://purl.org/dc/elements/1.1/}identifier').text
                identifier = identifier.replace(':','_')
                #rename the OBJ file to original filename pulled from MODS.xml
                os.rename(path_objfilename, os.path.join(path_bagsdirdirectory, identifier + '.' + extension))
            elif re.search('^FULL_TEXT', file):
                obj_file_name = file
                extension = obj_file_name.split('.')[1].strip()
                path_objfilename = os.path.join(proj_path, container, directory, obj_file_name)
                #use xml.etree to identify filename from MODS.xml
                tree = ET.parse(os.path.join(path_bagsdirdirectory, 'DC.xml'))
                identifier = tree.find('{http://purl.org/dc/elements/1.1/}identifier').text
                identifier = identifier.replace(':','_')
                #rename the OBJ file to original filename pulled from MODS.xml
                os.rename(path_objfilename, os.path.join(path_bagsdirdirectory, identifier + '.' + extension))
            elif re.search('^MP4', file):
                obj_file_name = file
                extension = obj_file_name.split('.')[1].strip()
                path_objfilename = os.path.join(proj_path, container, directory, obj_file_name)
                #use xml.etree to identify filename from MODS.xml
                tree = ET.parse(os.path.join(path_bagsdirdirectory, 'DC.xml'))
                identifier = tree.find('{http://purl.org/dc/elements/1.1/}identifier').text
                identifier = identifier.replace(':','_')
                #rename the OBJ file to original filename pulled from MODS.xml
                os.rename(path_objfilename, os.path.join(path_bagsdirdirectory, identifier + '.' + extension))
        num_bags += 1
    print('Processed {} bags'.format(str(num_bags)))

#this function continues to create the PAX structure by creating a "Representation_Access" folder and creating individual subdirs for all access assets in it
def representation_access():
    print('----CREATING REPRESENTATION_ACCESS FOLDERS----')
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    project_log_hand.close()
    container = vars[1].strip()
    folder_count = 0
    rep_acc = 'Representation_Access'
    path_container = os.path.join(proj_path, container)
    for directory in os.listdir(path = path_container):
        path_diracc = os.path.join(proj_path, container, directory, rep_acc)
        if directory.startswith('bags_'):
            print('bags_ folder found - skipped')
        else:
            os.mkdir(path_diracc)
            print('created {}'.format(path_diracc))
        folder_count += 1
    print('Created {} Representation_Access directories'.format(folder_count))

#this function creates an "access_ids.txt" file to store an identifier pulled from the MODS record as well as the path to the access assets in "bags_dir"
#this will then allow the access assets to be merged into the "container" folder in their "Representation_Access" subdirectories
def access_id_path():
    print('----CREATING LOG OF IDENTIFIERS AND FILE PATHS----')
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    project_log_hand.close()
    container = vars[1].strip()
    bags_dir = vars[2].strip()
    access_id_hand = open(os.path.join(proj_path, 'access_ids.txt'), 'a')
    access_count = 0
    path_bagsdir = os.path.join(proj_path, container, bags_dir)
    for directory in os.listdir(path = path_bagsdir):
        path_bagsdirdirectory = os.path.join(proj_path, container, bags_dir, directory)
        tree = ET.parse(os.path.join(path_bagsdirdirectory, 'MODS.xml'))
        identifier = tree.find('{http://www.loc.gov/mods/v3}identifier').text
        access_id_hand.write(identifier + '|')
        access_id_hand.write(path_bagsdirdirectory + '\n')
        access_count += 1
        print('logged {} and {}'.format(identifier, path_bagsdirdirectory))
    print('Logged {} paths and identifiers in access_ids.txt'.format(access_count))
    access_id_hand.close()

#this function loops through each subdir inside "container" and in each loop then loops through the entirety of "access_ids.txt" looking for a match
#between the MODS identifier recorded and checking to see if it matches the subdir name. If it does, it moved the access assets and metadata over
def merge_access_preservation():
    print('----MERGING ACCESS AND PRESERVATION ASSETS----')
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    project_log_hand.close()
    container = vars[1].strip()
    access_id_hand = open(os.path.join(proj_path, 'access_ids.txt'), 'r')
    access_id_list = access_id_hand.readlines()
    access_id_hand.close()
    rep_acc = 'Representation_Access'
    file_count = 0
    path_container = os.path.join(proj_path, container)
    for directory in os.listdir(path = path_container):
        path_directory = os.path.join(proj_path, container, directory)
        if directory.startswith('bags_'):
            continue
        else:
            for line in access_id_list:
                print('merging {} and {}'.format(directory,line))
                access_info = line.split('|')
                identifier = access_info[0].strip()
                path = access_info[1].strip()
                if identifier == directory:
                    for file in os.listdir(path = path):
                        if file.endswith('.xml'):
                            shutil.move(os.path.join(path, file), os.path.join(path_directory, file))
                            file_count += 1
                        else:
                            file_name = file.split('.')[0]
                            os.mkdir(os.path.join(path_directory, rep_acc, file_name))
                            shutil.move(os.path.join(path, file), os.path.join(path_directory, rep_acc, file_name, file))
                            file_count += 1
    print('Moved {} access and metadata files'.format(file_count))
    
#an alternative to the seperate functions that merge access and representation copies
#if bags contain both preservation and access copies, this function can structure the PAX representation folders in one function
#this specific instance also assumes two different types of materials (textual and video) where the textual has two different
#access copies and zero preservation representations
def representation_preservation_access():
    print('----CREATING REPRESENTATION_ACCESS FOLDERS----')
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    project_log_hand.close()
    container = vars[1].strip()
    folder_count = 0
    file_count = 0
    rep_acc = 'Representation_Access'
    rep_acc1 = 'Representation_Access_1'
    rep_acc2 = 'Representation_Access_2'
    rep_pres = 'Representation_Preservation'
    path_container = os.path.join(proj_path, container)
    for directory in os.listdir(path = path_container):
        path_directory = os.path.join(proj_path, container, directory)
        file_list = []
        for file in os.listdir(path = path_directory):
            file_list.append(file)
        testvartext = directory + '.txt'
        testvarvideo = directory + '.mp4'
        if testvartext in file_list:
            path_diracc1 = os.path.join(proj_path, container, directory, rep_acc1)
            os.mkdir(path_diracc1)
            path_diracc1_subdir = os.path.join(proj_path, container, directory, rep_acc1, directory)
            os.mkdir(path_diracc1_subdir)
            path_diracc2 = os.path.join(proj_path, container, directory, rep_acc2)
            os.mkdir(path_diracc2)
            path_diracc2_subdir = os.path.join(proj_path, container, directory, rep_acc2, directory)
            os.mkdir(path_diracc2_subdir)
            for file in os.listdir(path = path_directory):
                if file.endswith('pdf'):
                    shutil.move(os.path.join(path_directory, file), os.path.join(path_diracc1_subdir, file))
                    file_count += 1
                elif file.endswith('txt'):
                    shutil.move(os.path.join(path_directory, file), os.path.join(path_diracc2_subdir, file))
                    file_count += 1
            print('created {} and {}'.format(path_diracc1_subdir, path_diracc2_subdir))
        elif testvarvideo in file_list:
            path_diracc = os.path.join(proj_path, container, directory, rep_acc)
            os.mkdir(path_diracc)
            path_diracc_subdir = os.path.join(proj_path, container, directory, rep_acc, directory)
            os.mkdir(path_diracc_subdir)
            path_dirpres = os.path.join(proj_path, container, directory, rep_pres)
            os.mkdir(path_dirpres)
            path_dirpres_subdir = os.path.join(proj_path, container, directory, rep_pres, directory)
            os.mkdir(path_dirpres_subdir)
            for file in os.listdir(path = path_directory):
                if file.endswith('mp4'):
                    shutil.move(os.path.join(path_directory, file), os.path.join(path_diracc_subdir, file))
                    file_count += 1
                elif file.endswith('mov'):
                    shutil.move(os.path.join(path_directory, file), os.path.join(path_dirpres_subdir, file))
                    file_count += 1
            print('created {} and {}'.format(path_diracc_subdir, path_dirpres_subdir))
        folder_count += 1
    print('Created {} Representation Access or Preservation directories and moved {} files'.format(folder_count, file_count))

#this funciton simply removes the "bags_dir" folder path as well as deleting the "access_ids.txt" file
def cleanup_bags():
    print('----CLEANING UP BAGS----')
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    project_log_hand.close()
    container = vars[1].strip()
    bags_dir = vars[2].strip()
    shutil.rmtree(os.path.join(proj_path, container, bags_dir))
    os.remove('access_ids.txt')
    print('Deleted "{}" directory and access_ids.txt'.format(bags_dir))

#this function uses regex to remove the XML header from any metadata files before they are merged into a single OPEX file
#extra XML headers will cause the OPEX Incremental Workflow to fail when trying to ingest
def cleanup_metadata():
    print('----REMOVING EXTRA XML HEADERS----')
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    project_log_hand.close()
    container = vars[1].strip()
    header_count = 0
    path_container = os.path.join(proj_path, container)
    for directory in os.listdir(path = path_container):
        path_directory = os.path.join(proj_path, container, directory)
        for file in os.listdir(path = path_directory):
            if file.endswith('.xml'):
                temp_hand = open(os.path.join(path_directory, file), 'r')
                md_file = temp_hand.read()
                temp_hand.close()
                xml_header = re.findall('<\?.+\?>', md_file)
                if len(xml_header) > 0:
                    xml_header = xml_header[0]
                    new_md_file = md_file.replace(xml_header, '')
                    temp_hand = open(os.path.join(path_directory, file), 'w')
                    temp_hand.write(new_md_file)
                    temp_hand.close()
                    header_count += 1
                    print('removing XML header from {} in {}'.format(file, directory))
    print('Removed {} extra XML headers from metadata files'.format(header_count))
    
#this function creates the OPEX metadata file that accompanies an individual zipped PAX package
#this includes all the identifiers from the DC metadata file as well as the full MODS and DC records themselves
#this function also includes the metadata necessary for ArchivesSpace sync to Preservica
def pax_metadata():
    print('---CREATING METADATA FILES FOR PAX OBJECTS----')
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    project_log_hand.close()
    container = vars[1].strip()
    dir_count = 0
    path_container = os.path.join(proj_path, container)
    for directory in os.listdir(path = path_container):
        path_directory = os.path.join(proj_path, container, directory)
        try:
            opex1 = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><opex:OPEXMetadata xmlns:opex="http://www.openpreservationexchange.org/opex/v1.0"><opex:Properties><opex:Title>'
            tree = ET.parse(os.path.join(path_directory, 'DC.xml'))
            root = tree.getroot()
            opex2 = tree.find('{http://purl.org/dc/elements/1.1/}title').text
            opex3 = '</opex:Title><opex:Identifiers>'
            id_list = []
            opex4 = ''
            for id in root.findall('{http://purl.org/dc/elements/1.1/}identifier'):
                id_list.append(id.text)
            for item in id_list:
                if item.startswith('ur'):
                    opex4 += '<opex:Identifier type="code">' + item + '</opex:Identifier>'
                else:
                    other_identifiers = item.split(':')
                    label = other_identifiers[0].strip()
                    value = other_identifiers[1].strip()
                    opex4 += '<opex:Identifier type="' + label + '">' + value + '</opex:Identifier>'
            opex5 = '</opex:Identifiers></opex:Properties><opex:DescriptiveMetadata><LegacyXIP xmlns="http://preservica.com/LegacyXIP"><AccessionRef>catalogue</AccessionRef></LegacyXIP>'
            opex6 = ''
            for file in os.listdir(path = path_directory):
                if file.endswith('.xml'):
                    temp_file_hand = open(os.path.join(path_directory, file), 'r')
                    metadata = temp_file_hand.read()
                    metadata = metadata.strip()
                    opex6 += metadata + '\n'
                    temp_file_hand.close()
            opex7 = '</opex:DescriptiveMetadata></opex:OPEXMetadata>'
            filename = directory + '.pax.zip.opex'
            pax_md_hand = open(os.path.join(path_directory, filename), 'a')
            pax_md_hand.write(opex1 + opex2 + opex3 + opex4 + opex5 + opex6 + opex7)
            pax_md_hand.close()
            print('created {}'.format(filename))
            dir_count += 1
        except:
            print('ERROR: {}'.format(directory))
    print('Created {} OPEX metdata files for individual assets'.format(dir_count))

#this function stages the "Representation_Access" and "Representation_Preservation" folders for each asset inside a new directory
#this facilitates the creation of the zipped PAX package in the following function
def stage_pax_content():
    print('----STAGING PAX CONTENT IN PAX_STAGE----')
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    container = vars[1].strip()
    project_log_hand.close()
    pax_count = 0
    rep_count = 0
    path_container = os.path.join(proj_path, container)
    for directory in os.listdir(path = path_container):
        path_directory = os.path.join(proj_path, container, directory)
        path_paxstage = os.path.join(proj_path, container, directory, 'pax_stage')
        os.mkdir(path_paxstage)
        pax_count += 1
        shutil.move(os.path.join(path_directory, 'Representation_Access'), path_paxstage)
        shutil.move(os.path.join(path_directory, 'Representation_Preservation'), path_paxstage)
        rep_count += 2
        print('created /pax_stage in {}'.format(directory))
    print('Created {} pax_stage subdirectories and staged {} representation subdirectories'.format(pax_count, rep_count))

#this function takes the contents of the "pax_stage" folder created in the previous function and writes them into a zip archive
#the zip archive is the PAX object that will eventually become an Asset in Preservica
def create_pax():
    print('----CREATING PAX ZIP ARCHIVES----')
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    container = vars[1].strip()
    project_log_hand.close()
    dir_count = 0
    path_container = os.path.join(proj_path, container)
    for directory in os.listdir(path = path_container):
        path_zipdir = os.path.join(proj_path, container, directory, 'pax_stage/')
        path_directory = os.path.join(proj_path, container, directory)
        zip_dir = pathlib.Path(path_zipdir)
        pax_obj = ZipFile(os.path.join(path_directory, directory + '.zip'), 'w')
        for file_path in zip_dir.rglob("*"):
            pax_obj.write(file_path, arcname = file_path.relative_to(zip_dir))
        pax_obj.close()
        os.rename(os.path.join(path_directory, directory + '.zip'), os.path.join(path_directory, directory + '.pax.zip'))
        dir_count += 1
        print('created {}'.format(str(dir_count) + ': ' + directory + '.pax.zip'))
    print('Created {} PAX archives for ingest'.format(dir_count))

#this function deletes many files and folders that have now served their purpose in the migration process
#all metadata files are deleted as well as the "pax_stage" folder and it's contents
#a warning is thrown up and directory and file name information written to "project_log.txt" if an unexpected file is discovered
def cleanup_directories():
    print('----REMOVING UNNECESSARY FILES----')
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    project_log_hand.close()
    container = vars[1].strip()
    file_count = 0
    dir_count = 0
    unexpected = 0
    project_log_hand = open(proj_log_file, 'a')
    path_container = os.path.join(proj_path, container)
    for directory in os.listdir(path = path_container):
        path_directory = os.path.join(proj_path, container, directory)
        for entity in os.listdir(path = path_directory):
            path_entity = os.path.join(proj_path, container, directory, entity)
            if entity.endswith('.zip') == True:
                print('PAX: ' + entity)
            elif entity.endswith('.opex') == True:
                print('metadata: ' + entity)
            elif entity.endswith('.xml') == True:
                os.remove(path_entity)
                file_count += 1
                print('removed metadata file')
            elif os.path.isdir(path_entity) == True:
                shutil.rmtree(path_entity)
                dir_count += 1
                print('removed pax_stage directory')
            else:
                print('***UNEXPECTED ENTITY: ' + entity)
                project_log_hand.write('Unexpected entity in cleanup_directories(): ',directory,' | ',entity)
                unexpected += 1
    print('Deleted {} metadata files and {} Representation_Preservation and Representation_Access folders'.format(file_count, dir_count))
    print('Found {} unexpected entities'.format(unexpected))
    project_log_hand.close()

#this function loops through ever directory in "container" and opens up the OPEX metadata for the asset storing the entire contents in a string variable
#then the function loops through a text file that was manually created, containing the call number identifier as well as the ArchivesSpace archival object number
#while looping through the text file, if a match is found the OPEX metadata string variable, a metadata file is created for the folder
#this metadata is another facet required for ArchivesSpace to Preservica synchronization
def ao_opex_metadata():
    print('----CREATE ARCHIVAL OBJECT OPEX METADATA----')
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    container = vars[1].strip()
    project_log_hand.close()
    file_count = 0
    id_hand = open(os.path.join(proj_path, 'perkins-gillman_aonum_islid.txt'), 'r')
    id_list = id_hand.readlines()
    id_hand.close()
    path_container = os.path.join(proj_path, container)
    for directory in os.listdir(path = path_container):
        path_directory = os.path.join(proj_path, container, directory)
        if directory.startswith('archival_object_'):
            continue
        else:
            try:
                opex_hand = open(os.path.join(path_directory, directory + '.pax.zip.opex'), 'r')
                opex_str = opex_hand.read()
                opex_hand.close()
                ao_num = ''
                for line in id_list:
                    ids = line.split('|')
                    aonum = ids[0].strip()
                    isnum = ids[1].strip()
                    if opex_str.find(isnum) != -1:
                        ao_num = aonum
                        print('found a match for {} and {}'.format(aonum, isnum))
                opex = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><opex:OPEXMetadata xmlns:opex="http://www.openpreservationexchange.org/opex/v1.0"><opex:Properties><opex:Title>' + ao_num + '</opex:Title><opex:Identifiers><opex:Identifier type="code">' + ao_num + '</opex:Identifier></opex:Identifiers></opex:Properties><opex:DescriptiveMetadata><LegacyXIP xmlns="http://preservica.com/LegacyXIP"><Virtual>false</Virtual></LegacyXIP></opex:DescriptiveMetadata></opex:OPEXMetadata>'
                ao_md_hand = open(os.path.join(path_directory, ao_num + '.opex'), 'w')
                ao_md_hand.write(opex)
                ao_md_hand.close()
                os.rename(path_directory, os.path.join(proj_path, container, ao_num))
                file_count += 1
            except:
                print('error: {}'.format(directory))
    print('Created {} archival object metadata files'.format(file_count))

#this function creates the last OPEX metadata file required for the OPEX incremental ingest, for the container folder
#this OPEX file has the folder manifest to ensure that content is ingested properly
def write_opex_container_md():
    print('----CREATE CONTAINER OBJECT OPEX METADATA----')
    project_log_hand = open(proj_log_file, 'r')
    vars = project_log_hand.readlines()
    project_log_hand.close()
    container = vars[1].strip()
    opex1 = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><opex:OPEXMetadata xmlns:opex="http://www.openpreservationexchange.org/opex/v1.0"><opex:Transfer><opex:Manifest><opex:Folders>'
    opex2 = ''
    path_container = os.path.join(proj_path, container)
    for directory in os.listdir(path = path_container):
        opex2 += '<opex:Folder>' + directory + '</opex:Folder>'
    opex3 = '</opex:Folders></opex:Manifest></opex:Transfer></opex:OPEXMetadata>'
    container_opex_hand = open(os.path.join(proj_path, container, container + '.opex'), 'w')
    container_opex_hand.write(opex1 + opex2 + opex3)
    print('Created OPEX metadata file for {} directory'.format(container))
    container_opex_hand.close()

# create_container()
# folder_ds_files()
# create_bags_dir()
# extract_bags()
# validate_bags()
# revert_bags()
# rename_bags()
# create_id_ss()
# representation_preservation_access()
# representation_preservation()
# process_bags()
# representation_access()
# access_id_path()
# merge_access_preservation()
# cleanup_bags()
# pax_metadata()
# stage_pax_content()
# create_pax()
# cleanup_directories()
# ao_opex_metadata()
# write_opex_container_md()
